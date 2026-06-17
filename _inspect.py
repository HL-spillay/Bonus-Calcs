import openpyxl, sys

files = {
    "AUDIT":   "Formula extraction 2026-05 sales targets for communication_Audit.xlsx",
    "FORMULA": "Formula extraction 2026-05 sales targets for communication_Formula.xlsx",
}

for tag, path in files.items():
    print("=" * 70)
    print(tag, "->", path)
    print("=" * 70)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print("  ERROR opening:", e)
        continue
    for ws in wb.worksheets:
        print(f"  - {ws.title!r:50}  max_row={ws.max_row} max_col={ws.max_column}")
    wb.close()
    print()
