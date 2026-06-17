import openpyxl
from openpyxl.utils import get_column_letter

PATH = "Formula extraction 2026-05 sales targets for communication_Audit.xlsx"
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

for name, maxr, maxc in [("Managers criteria", 140, 6), ("Store Staff Criteria", 40, 7)]:
    print("=" * 80)
    print(name)
    print("=" * 80)
    ws = wb[name]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > maxr:
            break
        cells = []
        for j, v in enumerate(row[:maxc]):
            if v is None or str(v).strip() == "":
                continue
            cells.append(f"{get_column_letter(j+1)}{i}={v}")
        if cells:
            print("  " + " | ".join(cells))
    print()
wb.close()
