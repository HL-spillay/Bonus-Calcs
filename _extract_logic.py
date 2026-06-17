import openpyxl, json
from openpyxl.utils import get_column_letter

FPATH = "Formula extraction 2026-05 sales targets for communication_Formula.xlsx"
wb = openpyxl.load_workbook(FPATH, read_only=True, data_only=True)

names = wb.sheetnames
formula_tabs = [n for n in names if n.startswith("Formulas - ")]
original = [n for n in names if not n.startswith("Formulas - ")]

def match_formula_tab(orig):
    # formula tab names are truncated; match by prefix of remainder
    cands = []
    for ft in formula_tabs:
        rem = ft[len("Formulas - "):]
        if orig.startswith(rem) or rem.startswith(orig[:len(rem)]):
            cands.append((len(rem), ft))
    if not cands:
        return None
    cands.sort(reverse=True)
    return cands[0][1]

def read_grid(ws, max_rows):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append(row)
    return rows

result = {}
for orig in original:
    if orig in ("Execution Log",):
        continue
    ows = wb[orig]
    orig_rows = read_grid(ows, 4)            # headers + a few sample values
    headers = orig_rows[0] if orig_rows else ()
    sample = orig_rows[1] if len(orig_rows) > 1 else ()

    ft = match_formula_tab(orig)
    col_formulas = {}
    if ft:
        fws = wb[ft]
        frows = read_grid(fws, 30)           # scan first 30 rows for a representative formula per col
        ncols = max((len(r) for r in frows), default=0)
        for c in range(ncols):
            rep = ""
            rep_row = None
            for ri, r in enumerate(frows[1:], start=2):  # skip header row
                if c < len(r) and r[c] and str(r[c]).startswith("="):
                    rep = str(r[c])
                    rep_row = ri
                    break
            if rep:
                col_formulas[c] = (rep_row, rep)

    result[orig] = {
        "formula_tab": ft,
        "headers": [("" if h is None else str(h)) for h in headers],
        "sample": [("" if s is None else str(s)) for s in sample],
        "col_formulas": {str(k): v for k, v in col_formulas.items()},
    }

wb.close()

with open("_logic_dump.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=1)

# also a human-readable text version
with open("_logic_dump.txt", "w", encoding="utf-8") as f:
    for sheet, d in result.items():
        f.write("=" * 90 + "\n")
        f.write(f"SHEET: {sheet}   (formula tab: {d['formula_tab']})\n")
        f.write("=" * 90 + "\n")
        hdr = d["headers"]
        smp = d["sample"]
        for c in range(max(len(hdr), len(smp), max((int(k) for k in d['col_formulas']), default=-1) + 1)):
            letter = get_column_letter(c + 1)
            h = hdr[c] if c < len(hdr) else ""
            s = smp[c] if c < len(smp) else ""
            cf = d["col_formulas"].get(str(c))
            line = f"  {letter:>3} | hdr={h[:32]:32} | sample={s[:24]:24}"
            if cf:
                line += f" | F[r{cf[0]}]={cf[1][:240]}"
            f.write(line + "\n")
        f.write("\n")

print("wrote _logic_dump.json and _logic_dump.txt")
print("sheets processed:", len(result))
